import pandas as pd
from datetime import datetime, timedelta
from pytz import timezone
from crosschex_cloud_api import get_crosschex_token, get_all_records
from dateutil import parser
import numpy as np
import time

pd.options.mode.chained_assignment = None

# from enum import IntEnum

# Constants
T0 = datetime.strptime("2024-02-19", "%Y-%m-%d").astimezone(timezone("US/Eastern"))
PAY_PERIOD_LENGTH = 14
DAY_OF_WEEK = {0: "Mon", 1: "Tue", 2: "Wed", 3: "Thu", 4: "Fri", 5: "Sat", 6: "Sun"}


"""
1: 'IT and Inventory Manager',
2:	'Assistant Manager',
4:	'OBL Manager',
5:	'MA',
6:	'RN',
7:	'Office Assistant',
8:	'Patient Intake Specialist',
9:	'Sonographer',

"""
employee_postion_dict = {
    1: "IT and Inventory Manager",
    2: "Assistant Manager",
    4: "OBL Manager",
    5: "MA",
    6: "RN",
    7: "Office Assistant",
    8: "Patient Intake Specialist",
    9: "Sonographer",
    10: "MA",
    11: "Call Center Team",
    12: "MA",
}

empl_id_name_dict = {
    1: "Yousaf Chaudry",
    2: "Mohamad Majdalawi",
    4: "Margaret Kline",
    5: "Hannah Billings",
    6: "Alex Wells",
    7: "Victoria Ulrich",
    8: "Ahmed Chaudry",
    9: "Tamrat Hailemariam",
    10: "Justine Elgen",
    11: "Maya Wells",
    12: "Eliza Aamir",
}


"""
from crosschex_cloud_api import get_all_records, get_crosschex_token
from time_parse import process_json_response, processing
token = get_crosschex_token()
response = get_all_records(token)
df = process_json_response(response)
r = processing(df)
--- then run prepare_timesheet_data(r)

05/01/2024:
if need all days an employee worked, then pass int of employee id to the get_previous_pp method of crosschex_cloud_api,
this will pull specific employee records only. 

02/29/2024: future implementation of creating lists and then grouped by year, of all future payperiods from T0 until any given date:

future_date = datetime(2030, 6, 25)
last_pp = calculate_pay_period(future_date)
periods = []
for n in range(last_pp[0]):
	date = T0 (which is 2-19-2024)  + timedelta(days = 14 * n)
	periods.append(calculate_pay_period(date))

df9 = pd.DataFrame(periods)
df9.rename(columns={0:"pay_period", 1:"start_date", 2:"end_date"}, inplace=True)
df9['year_of_payperiod'] = df9['start_date'].apply(lambda row: row.year)
c = z.groupby(['year_of_payperiod']) 


"""


# Functions


def calculate_hours_from_str(a, b):
    from dateutil.tz import gettz

    tzinfos = {}
    start = parser.parse(a)
    end = parser.parse(b)
    if start > end:
        start, end = end, start
    duration = end - start
    duration_hours = duration.total_seconds() / 3600
    return duration_hours


def calculate_pay_period(date):
    delta = date.astimezone(timezone("US/Eastern")) - T0
    period_number = delta.days // PAY_PERIOD_LENGTH
    period_start = T0 + timedelta(days=period_number * PAY_PERIOD_LENGTH)
    period_end = period_start + timedelta(days=PAY_PERIOD_LENGTH - 1)
    return period_number, period_start, period_end


# def calculate_specific_year_pay_period(date):
#     current_year = datetime.now().year
#     current_years_t0 = (
#         current_year + "-01-01" + "is this correct? need to determine every year's T0"
#     )
#     delta = date.astimezone(timezone("US/Eastern")) - T0
#     period_number = delta.days // PAY_PERIOD_LENGTH
#     period_start = T0 + timedelta(days=period_number * PAY_PERIOD_LENGTH)
#     period_end = period_start + timedelta(days=PAY_PERIOD_LENGTH - 1)
#     return period_number, period_start, period_end


# 02/23/2024: need to modify this, or condense into .apply() as in calculate_pay_period
# ---- this function specifically if overlap occurs between shifts belonging to different periods
# ---- ie: 11pm to 7am shift, 11-12am will get counted to previous shift/payperiod, and 12am-7am to
# ---- following pay period/shift.
# def assign_shift_to_period(clock_in, clock_out):
#     _, period_start_in, period_end_in = calculate_pay_period(clock_in)
#     _, period_start_out, period_end_out = calculate_pay_period(clock_out)

#     if period_start_in == period_start_out or (
#         clock_in < period_start_out and clock_out <= period_end_out
#     ):
#         return period_start_out, period_end_out
#     else:
#         return period_start_out, period_end_out


# def calculate_shift_hours(clock_in, clock_out):
#     duration = clock_out - clock_in
#     hours = duration.total_seconds() / 3600
#     return hours


def is_leap_year(year):
    """
    Determine if a given year is a leap year.
    """
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def group_pay_periods_by_year(pay_periods):
    """
    Group pay periods into years based on the provided logic.
    """
    year_groups = {}
    for period, details in pay_periods.items():
        start_date = details["start"]
        end_date = details["end"]
        if start_date.year == end_date.year:
            year = start_date.year
        elif start_date.month == 12 and end_date.month == 1 and end_date.day <= 14:
            year = end_date.year
        else:
            year = start_date.year

        if year not in year_groups:
            year_groups[year] = []
        year_groups[year].append(period)

    return year_groups


def process_shifts(df):
    """
    02/23/2024: need to update this function, specifically 'shift_end' included inside of "paired" column,
    and also for df['pay_period'], it will have to be based on "date" column,
    and refactor "hours_worked" by using "hours_worked" column and just unpacking
    and subtracting smaller values from larger if applicable (ie breaks' values from the main shift)
    """

    # Sort and calculate shift durations
    df.sort_values(by=["workno", "checktime"], inplace=True)

    # Calculate pay period for each shift
    df["pay_period"] = df["checktime"].apply(lambda x: calculate_pay_period(x)[0])

    # Calculate hours worked for each shift assuming consecutive logins as in/out pairs
    df["shift_end"] = df.groupby("workno")["checktime"].shift(-1)
    df["hours_worked"] = (df["shift_end"] - df["checktime"]).dt.total_seconds() / 3600

    # Drop NaN values in 'shift_end' to filter out last entry or unmatched shifts
    df.dropna(subset=["shift_end"], inplace=True)

    # Group by pay period and sum hours worked
    pay_period_summary = (
        df.groupby(["pay_period", "workno"])["hours_worked"].sum().reset_index()
    )

    return pay_period_summary


def pay_periods_list():
    current_year = datetime.now().year
    last_day_of_year = str(current_year) + "-12-31"
    T0 = datetime.strptime("2024-02-19", "%Y-%m-%d").astimezone(timezone("US/Eastern"))
    if T0.year != current_year:
        print("time to change year in this code(reset T0)")
        raise ValueError
        # last_pay_period_current_year = calculate_pay_period(
        #     datetime.strptime(last_day_of_year, "%Y-%m-%d")
        # )
        # # first_pay_period = calculate_pay_period(T0)
        # last_day_of_t0 = T0.year + "-12-31"
        # last_pay_period_t0_year = calculate_pay_period(
        #     datetime.strptime(last_day_of_t0, "%Y-%m-%d")
        # )
        # periods = []
        # total_periods_elapsed = (
        #     last_pay_period_current_year[0] + last_pay_period_t0_year[0]
        # )
        # for n in range(total_periods_elapsed):
        #     date = T0 + timedelta(days=14 * n)
        #     periods.append(calculate_pay_period(date))
        # return periods
    last_pay_period = calculate_pay_period(
        datetime.strptime(last_day_of_year, "%Y-%m-%d")
    )
    # first_pay_period = calculate_pay_period(T0)
    periods = []
    for n in range(last_pay_period[0]):
        date = T0 + timedelta(days=14 * n)
        periods.append(calculate_pay_period(date))
    return periods


def process_json_response(response):
    data_list = []
    now = datetime.now().strftime("%m_%d_%Y_%H%M")
    for i in response["payload"][
        "list"
    ]:  # this is section where each entry of clockin occurs
        first_name = i["employee"]["first_name"]
        last_name = i["employee"]["last_name"]
        workno = i["employee"]["workno"]
        dept = i["employee"]["department"]
        checktime = i["checktime"]
        job_title = i["employee"]["job_title"]
        device = i["device"]["name"]
        data_list.append(
            {
                "first_name": first_name,
                "last_name": last_name,
                "workno": workno,
                "dept": dept,
                "checktime": checktime,
                "job_title": job_title,
                "device": device,
            }
        )
    df_main = pd.DataFrame(data_list)
    # df_main["checktime"] = pd.to_datetime(df_main["checktime"])
    # df_main["checktime"] = df_main["checktime"].apply(
    #     lambda x: x.astimezone(timezone("US/Eastern"))
    # )
    # df_main.to_csv(f"processed_json_df{now}.csv", index=False)
    return df_main
    # summary_df = process_shifts(df_main)
    # summary_df.to_csv("latest_test.csv")
    # return summary_df


def calc_hours_minus_breaks(times):
    if (
        len(times) < 2
        or not isinstance(times, list)
        or not all([t.isdigit() for t in times if type(t) not in [int, float]])
    ):
        if (
            isinstance(times, list)
            and len(times) == 1
            and type(times[0]) in [int, float]
        ):
            (val,) = times
            return round(val, 2)
        return times
    nums = [float(time) for time in times]
    highest_val = max(nums)
    nums.remove(highest_val)
    result = highest_val
    for n in nums:
        result = result - n
    return round(result, 2)


def processing(data):
    """02/23/2024: passing group['checktime'].iloc[0].date().weekday() to DAY_OF_WEEK dict, yields 3 letter abbrev."""
    df = process_json_response(data) if not isinstance(data, pd.DataFrame) else data
    # df = process_json_response(data)
    employee_dfs = []
    # summary_df = pd.DataFrame(
    #     columns=[
    #         "Pay_Period",
    #         "Employee_ID",
    #         "Employee_Name",
    #         "Dept",
    #         "Total_Hours_Worked",
    #     ]
    # )
    import time

    if time.daylight:
        delta = timedelta(hours=-1)
        df["backup_checktime"] = df["checktime"]
        df["checktime"] = pd.to_datetime(df["checktime"]).apply(lambda x: x + delta)
        df["checktime"] = df["checktime"].dt.tz_convert("US/Eastern")
    else:
        df["checktime"] = pd.to_datetime(df["checktime"]).dt.tz_convert("US/Eastern")
    # df["checktime"] = df["checktime"].apply(
    #     lambda x: x.astimezone(timezone("US/Eastern"))
    # )
    employee_dfs = [employee_group for _, employee_group in df.groupby("workno")]
    all_employees_sorted_df = pd.DataFrame()

    for employee in employee_dfs:
        sorted_employee_df = pd.DataFrame()
        grouped_by_date = employee.groupby(employee["checktime"].dt.date)

        for date, group in grouped_by_date:
            group.sort_values(by=["checktime"], inplace=True)
            as_list = group["checktime"].tolist()
            arr = np.array(as_list)
            paired = (
                [(arr[i], arr[-i - 1]) for i in range(len(arr) // 2)]
                if len(as_list) % 2 == 0
                else [arr[i] for i in range(len(arr))]
            )
            group["date"] = date.strftime("%m-%d-%Y")
            (
                group["pay_period"],
                group["paired"],
                group["hours_worked"],
                group["paired_data"],
                group["day"],
                # group["pp_hours_summary"],
            ) = (None, None, None, None, None)
            group["paired"].iloc[0] = (
                [(f"{x[0].strftime('%H%M')}-{x[1].strftime('%H%M')}") for x in paired]
                if len(as_list) % 2 == 0
                else [x.strftime("%H%M") for x in paired]
            )
            group.dropna(subset=["paired"], inplace=True)
            hours_data = (
                [
                    calculate_hours_from_str(
                        x[0].strftime("%m/%d/%Y %H%M"), x[1].strftime("%m/%d/%Y %H%M")
                    )
                    for x in paired
                ]
                if len(as_list) % 2 == 0
                else f"Manually calculate, data missing/incorrect; notify {group['first_name'].iloc[0]}, Employee ID-{group['workno'].iloc[0]}, to file Attendance Request for {group['date'].iloc[0]}."
            )
            # passing group['checktime'].iloc[0].date().weekday() to DAY_OF_WEEK dict, yields 3 letter abbrev."""
            day = group["checktime"].iloc[0].date().weekday()
            group["day"].iloc[0] = DAY_OF_WEEK[day]
            group["hours_worked"].iloc[0] = calc_hours_minus_breaks(hours_data)
            group["paired_data"].iloc[0] = paired
            group["device"].iloc[0] = (
                "APP" if group["device"].iloc[0] == "" else group["device"].iloc[0]
            )
            group["name"] = group["first_name"] + " " + group["last_name"]
            group["pay_period"] = group["checktime"].apply(
                lambda row: calculate_pay_period(row)[0]
            )
            group.rename(
                columns={
                    "checktime": "clock_entry",
                    "job_title": "role",
                    "workno": "employee_id",
                    "paired": "paired_clock_times",
                    "paired_data": "reference_data",
                    "pay_period": "pay_period_num",
                    "date": "shift_date",
                },
                inplace=True,
            )
            group = group.reindex(
                columns=[
                    # "pp_hours_summary",
                    "pay_period_num",
                    "day",
                    "shift_date",
                    "name",
                    "employee_id",
                    "hours_worked",
                    "device",
                    "paired_clock_times",
                    "role",
                    "dept",
                    "clock_entry",
                    "reference_data",
                ]
            )

            sorted_employee_df = pd.concat(
                [sorted_employee_df, group], axis=0, ignore_index=True
            )

        all_employees_sorted_df = pd.concat(
            [all_employees_sorted_df, sorted_employee_df], axis=0, ignore_index=True
        )
    all_employees_sorted_df["pay_period_dates"] = all_employees_sorted_df[
        "clock_entry"
    ].apply(lambda row: pd.Series([calculate_pay_period(row)]).to_numpy())
    all_employees_sorted_df["pp_dates"] = all_employees_sorted_df[
        "pay_period_dates"
    ].apply(
        lambda row: row[0][1].strftime("%m-%d-%Y")
        + " - "
        + row[0][2].strftime("%m-%d-%Y")
    )
    all_employees_sorted_df.loc[
        all_employees_sorted_df["hours_worked"].astype(str).str.contains("Manual"),
        "hours_worked",
    ] = 0
    # all_employees_sorted_df.reset_index(drop=True, inplace=True)
    # new_column = all_employees_sorted_df.groupby(
    #     ["pay_period_num", "pp_dates", "employee_id", "name", "dept"], as_index=False
    # )["hours_worked"].aggregate(["sum"])
    # new_column.reset_index(level=0, drop=True)
    # all_employees_sorted_df["pp_hours_summary"] = all_employees_sorted_df.groupby(
    #     ["pay_period_num", "pp_dates", "employee_id", "name", "dept"], as_index=False
    # )["hours_worked"].aggregate(["sum"])
    all_employees_sorted_df["pp_hours_summary"] = all_employees_sorted_df.groupby(
        ["pay_period_num", "pp_dates", "employee_id", "name", "dept"]
    )["hours_worked"].transform(lambda x: x.sum())
    # all_employees_sorted_df["lunch_break_hours"] = all_employees_sorted_df.groupby(
    #     ["pay_period_num", "pp_dates", "employee_id", "name", "dept"]
    # )["hours_worked"].transform(lambda x: x.sum())

    return all_employees_sorted_df


def process_exported_timesheet(data):
    df = pd.read_csv(data)
    employee_dfs = []
    df.rename(
        columns={
            "Name": "name",
            "Employee No.": "employee_id",
            "Department": "dept",
            "Date": "date",
            "Time": "time",
            "Device": "device",
        },
        inplace=True,
    )
    df["position"] = df["employee_id"].apply(lambda x: employee_postion_dict[x])

    def combine_time(row):
        return datetime.strptime(row["date"] + " " + row["time"], "%m/%d/%Y %I:%M %p")

    df["checktime"] = df.apply(lambda x: combine_time(x), axis=1)
    df["checktime"] = df["checktime"].apply(lambda x: x.tz_localize("US/Eastern"))
    # .dt.tz_localize('US/Pacific')
    employee_dfs = [employee_group for _, employee_group in df.groupby("employee_id")]

    all_employees_sorted_df = pd.DataFrame()

    for employee in employee_dfs:
        sorted_employee_df = pd.DataFrame()
        grouped_by_date = employee.groupby(employee["checktime"].dt.date)

        for date, group in grouped_by_date:
            group.sort_values(by=["checktime"], inplace=True)
            as_list = group["checktime"].tolist()
            arr = np.array(as_list)
            paired = (
                [(arr[i], arr[-i - 1]) for i in range(len(arr) // 2)]
                if len(as_list) % 2 == 0
                else [arr[i] for i in range(len(arr))]
            )
            group["date"] = date.strftime("%m-%d-%Y")
            (
                group["pay_period"],
                group["paired"],
                group["hours_worked"],
                group["paired_data"],
                group["day"],
                # group["pp_hours_summary"],
            ) = (None, None, None, None, None)
            group["paired"].iloc[0] = (
                [(f"{x[0].strftime('%H%M')}-{x[1].strftime('%H%M')}") for x in paired]
                if len(as_list) % 2 == 0
                else [x.strftime("%H%M") for x in paired]
            )
            group.dropna(subset=["paired"], inplace=True)
            hours_data = (
                [
                    calculate_hours_from_str(
                        x[0].strftime("%m/%d/%Y %H%M"), x[1].strftime("%m/%d/%Y %H%M")
                    )
                    for x in paired
                ]
                if len(as_list) % 2 == 0
                else f"Manually calculate, data missing/incorrect; notify {group['name'].iloc[0]}, Employee ID-{group['employee_id'].iloc[0]}, to file Attendance Request for {group['date'].iloc[0]}."
            )
            day = group["checktime"].iloc[0].date().weekday()
            group["day"].iloc[0] = DAY_OF_WEEK[day]
            group["hours_worked"].iloc[0] = calc_hours_minus_breaks(hours_data)
            group["paired_data"].iloc[0] = paired
            group["pay_period"] = group["checktime"].apply(
                lambda row: calculate_pay_period(row)[0]
            )
            group = group.reindex(
                columns=[
                    # "pp_hours_summary",
                    "pay_period",
                    "day",
                    "date",
                    "name",
                    "employee_id",
                    "hours_worked",
                    "device",
                    "paired_clock_times",
                    "position",
                    "dept",
                    "checktime",
                    "paired_data",
                ]
            )
            sorted_employee_df = pd.concat(
                [sorted_employee_df, group], axis=0, ignore_index=True
            )
        all_employees_sorted_df = pd.concat(
            [all_employees_sorted_df, sorted_employee_df], axis=0, ignore_index=True
        )
    all_employees_sorted_df["pay_period_dates"] = all_employees_sorted_df[
        "checktime"
    ].apply(lambda row: pd.Series([calculate_pay_period(row)]).to_numpy())
    all_employees_sorted_df["pp_dates"] = all_employees_sorted_df[
        "pay_period_dates"
    ].apply(
        lambda row: row[0][1].strftime("%m-%d-%Y")
        + " - "
        + row[0][2].strftime("%m-%d-%Y")
    )
    all_employees_sorted_df.loc[
        all_employees_sorted_df["hours_worked"].astype(str).str.contains("Manual"),
        "hours_worked",
    ] = 0
    all_employees_sorted_df["pp_hours_summary"] = all_employees_sorted_df.groupby(
        ["pay_period", "pp_dates", "employee_id", "name", "dept"]
    )["hours_worked"].transform(lambda x: x.sum())

    return all_employees_sorted_df

    # # Calculate pay period for each shift
    # df["pay_period"] = df["checktime"].apply(lambda x: calculate_pay_period(x)[0])

    # # Calculate hours worked for each shift assuming consecutive logins as in/out pairs
    # df["shift_end"] = df.groupby("workno")["checktime"].shift(-1)
    # df["hours_worked"] = (df["shift_end"] - df["checktime"]).dt.total_seconds() / 3600

    # # Drop NaN values in 'shift_end' to filter out last entry or unmatched shifts
    # df.dropna(subset=["shift_end"], inplace=True)

    # # Group by pay period and sum hours worked
    # pay_period_summary = (
    #     df.groupby(["pay_period", "workno"])["hours_worked"].sum().reset_index()
    # )

    # return pay_period_summary


def calc_pp_year(future_date, current_year: bool = False, current_pp: bool = False):
    date = (
        datetime.strptime(future_date, "%m-%d-%Y")
        if not isinstance(future_date, datetime)
        else future_date
    )
    last_pp = calculate_pay_period(date)
    periods = []
    for n in range(last_pp[0]):
        date = T0 + timedelta(days=14 * n)
        periods.append(calculate_pay_period(date))
    df = pd.DataFrame(periods)
    df.rename(columns={0: "pay_period", 1: "start_date", 2: "end_date"}, inplace=True)
    df["year_of_payperiod"] = df["start_date"].apply(lambda row: row.year)
    grouped = df.groupby(["year_of_payperiod"])
    if current_pp:
        # df_year =
        return df[
            (df["start_date"].dt.date < datetime.now().date())
            & (df["end_date"].dt.date > datetime.now().date())
        ]
    if current_year:
        return [(a[0], b) for a, b in grouped if a[0] == datetime.now().year]
    return [(a[0], b) for a, b in grouped]


def calc_lunch(df):
    multiplier = 0.5
    counter = 0
    for idx, row in df.iterrows():
        if row["hours_worked"] > 6.0:
            counter += 1
    return counter * multiplier


def prepare_timesheet_data(df_origin):
    # the df sent to this function should be from processing()
    groupby_obj = df_origin.groupby("name", group_keys=True)[
        ["shift_date", "hours_worked", "pp_hours_summary"]
    ]
    df = pd.DataFrame([(a, b) for a, b in groupby_obj])
    # removing index column from shift_date data in each of column 1's dataframes
    df[1] = df[1].apply(lambda x: x.set_index(x.columns[0], drop=True))
    df.rename(columns={0: "employee_name", 1: "hours_data"}, inplace=True)
    # assigning sum of pp hours to a separate column
    df["pp_summary"] = df.hours_data.apply(lambda x: x["pp_hours_summary"].iloc[0])
    # getting lunch hours ratio result
    df["lunch_hours"] = df["hours_data"].transform(lambda x: calc_lunch(x))
    df["hours_worked_minus_lunch"] = df["pp_summary"] - df["lunch_hours"]
    df["hours_data"] = df.hours_data.apply(
        lambda x: x.drop(["pp_hours_summary"], axis=1)
    )
    df["shift_date_hours_worked"] = df["hours_data"].transform(
        lambda x: x.to_string(header=False)
    )
    df["shift_date_hours_worked"] = df["shift_date_hours_worked"].apply(
        lambda x: x.strip("shift_date").strip()  # .replace("  ", "          ").strip()
    )
    now = datetime.now().strftime("%m_%d_%Y_%H%M")
    df.drop(columns=["hours_data"], axis=1, inplace=True)
    df.rename(
        columns={
            "employee_name": "Employee_Name",
            "pp_summary": "Sum_Hours_Worked",
            "lunch_hours": "Lunch_Hours",
            "hours_worked_minus_lunch": "Hours_Worked_Minus_Lunch",
            "shift_date_hours_worked": "Shift_Date__Hours_Per_Day",
        },
        inplace=True,
    )
    df.reindex(
        columns=[
            "Employee_Name",
            "Shift_Date__Hours_Per_Day",
            "Sum_Hours_Worked",
            "Lunch_Hours",
            "Hours_Worked_Minus_Lunch",
        ]
    )

    employees = df_origin["employee_id"].unique().tolist()
    zero_hour_employees = [
        e for e in empl_id_name_dict.keys() if str(e) not in employees
    ]
    zero_hour_employees_list = []
    for z in zero_hour_employees:
        new_row = {
            "Employee_Name": empl_id_name_dict[z],
            "Shift_Date__Hours_Per_Day": "N/A",
            "Sum_Hours_Worked": 0,
            "Lunch_Hours": 0,
            "Hours_Worked_Minus_Lunch": 0,
        }
        zero_hour_employees_list.append(new_row)
        # df.loc[len(df)] = new_row
        # df.index = df.index + 1
        # df = df.sort_index()

    zero_hour_employees_df = pd.DataFrame(zero_hour_employees_list)

    df = pd.concat([df, zero_hour_employees_df], axis=0, ignore_index=True)

    # df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})
    xlsx_filename = f"timesheet_{now}.xlsx"
    writer = pd.ExcelWriter(xlsx_filename, engine="xlsxwriter")
    # df_detailed = df.copy()

    with writer as writer:
        df.to_excel(
            writer, sheet_name="Sheet1", startrow=4, index=False
        )  # <<< notice the startrow here

    import openpyxl as op
    from openpyxl.styles import Font, Alignment, Border

    # import xlwings as xw
    # from openpyxl.utils import get_column_letter

    wb = op.load_workbook(xlsx_filename)
    ws = wb.active
    main_title = Font(bold=True, underline="single", size=16)
    column_row_font = Font(bold=True, underline="single")
    general_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cols = ["A", "B", "C", "D", "E"]
    for col in cols:
        col_id = ws.column_dimensions[col]
        col_id.alignment = general_align
    for row in ws["A5:E5"]:
        for cell in row:
            cell.font = column_row_font
            # cell.alignment = general_align
    cell_a2 = ws["A2"]
    cell_a2.value = "Start Date to End Date: "
    cell_b2 = ws["B2"]
    cell_b2.value = df_origin.iloc[0]["pp_dates"]
    cell_a1 = ws["A1"]
    cell_a1.value = "Pay Period Detailed Timesheet"
    cell_a1.font = main_title
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35

    # def auto_col_width(ws):
    #     # dims = {}
    #     # for row in ws.rows:
    #     #     for cell in row:
    #     #         if cell.value:
    #     #             dims[cell.column_letter] = max(
    #     #                 (dims.get(cell.column_letter, 0), len(str(cell.value).strip()))
    #     #             )
    #     # for col, value in dims.items():
    #     #     ws.column_dimensions[col].width = value * 1.2
    #     ws.column_dimensions["A"].width = 25
    #     ws.column_dimensions["B"].width = 25
    #     ws.column_dimensions["C"].width = 25
    #     ws.column_dimensions["D"].width = 35
    #     ws.column_dimensions["E"].width = 35
    #     print(f"adjusted columns: A, B, C, D, E")

    # for column in ws.columns:
    #     max_length = 0
    #     column_letter = column[0].column_letter
    #     for cell in column:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2) * 1.2
    #     ws.column_dimensions[column_letter].width = adjusted_width
    #     print(f"adjusted column: {column}")

    # auto_col_width(ws)

    def auto_row_height(ws):
        for row in ws.iter_rows(
            min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column
        ):
            max_height = 0
            for cell in row:
                try:
                    # Calculate the number of lines in the cell content
                    lines = str(cell.value).count("\n") + 1

                    # Determine the height required based on the number of lines
                    height = lines * 15  # Adjust the factor as needed

                    # Update max_height if the calculated height is greater
                    if height > max_height:
                        max_height = height
                except TypeError:
                    pass

            # Set the row height to accommodate the tallest cell in the row
            ws.row_dimensions[row[0].row].height = max_height
            print(f"adjusted row: {row}")

    auto_row_height(ws)
    xlsx_updated = xlsx_filename.replace("timesheet", "timesheet_rev_1")
    wb.save(xlsx_updated)
    time.sleep(1.5)
    df_payroll_init = pd.read_excel(xlsx_updated, engine="openpyxl", skiprows=4)
    df_payroll_final = pd.concat(
        [df_payroll_init["Employee_Name"], df_payroll_init["Hours_Worked_Minus_Lunch"]],
        axis=1,
    )
    df_payroll_final.rename(
        columns={
            "Employee_Name": "Employee Name",
            "Hours_Worked_Minus_Lunch": "Hours Worked",
        },
        inplace=True,
    )
    xlsx_payroll = f"payroll_summary_{now}.xlsx"
    writer_payroll = pd.ExcelWriter(xlsx_payroll, engine="xlsxwriter")
    with writer_payroll as writer_payroll:
        df_payroll_final.to_excel(
            writer_payroll, sheet_name="Sheet1", startrow=4, index=False
        )
    time.sleep(0.5)
    wb_payroll = op.load_workbook(xlsx_payroll)
    ws_payroll = wb_payroll.active
    cols_payroll = ["A", "B"]
    for col in cols_payroll:
        col_id_payroll = ws_payroll.column_dimensions[col]
        col_id_payroll.alignment = general_align
    for row in ws_payroll["A5:B5"]:
        for cell in row:
            cell.font = column_row_font
            # cell.alignment = general_align
    cell_a2_pp = ws_payroll["A2"]
    cell_a2_pp.value = "Start Date to End Date: "
    cell_b2_pp = ws_payroll["B2"]
    cell_b2_pp.value = df_origin.iloc[0]["pp_dates"]
    cell_a1_pp = ws_payroll["A1"]
    cell_a1_pp.value = "Pay Period Summary for Payroll"
    cell_a1_pp.font = main_title
    ws_payroll.column_dimensions["A"].width = 25
    ws_payroll.column_dimensions["B"].width = 35
    auto_row_height(ws_payroll)
    wb_payroll.save(xlsx_payroll)

    # path_for_final = (
    #     r"C:\Users\omniv\OneDrive\Documents\py-testing" + rf"{xlsx_updated}"
    # )
    # auto_fit_excel_columns_and_rows(path_for_final)

    # ----- 03/25/2024: need to add
    # workbook = writer.book
    # worksheet = writer.sheets['Sheet1']
    # worksheet.write(row, 0, 'Some Text')
    return df


"""
    # employees = df_origin["employee_id"].unique().tolist()
    # zero_hour_employees = [
    #     e for e in empl_id_name_dict.keys() if str(e) not in employees
    # ]
    # for z in zero_hour_employees:
    #     new_row = {
    #         "Employee_Name": empl_id_name_dict[z],
    #         "Shift_Date__Hours_Per_Day": "N/A",
    #         "Sum_Hours_Worked": 0,
    #         "Lunch_Hours": 0,
    #         "Hours_Worked_Minus_Lunch": 0,
    #     }
    #     df.loc[len(df)] = new_row
    #     df.index = df.index + 1
    #     df = df.sort_index()

"""
